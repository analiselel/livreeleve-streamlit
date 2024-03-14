import requests
from openpyxl import load_workbook
import time

def update_stock():
    # Carregar o arquivo Excel gerado pelo script anterior
    wb = load_workbook("dados.xlsx")
    ws = wb.active
    
    # Definir as informações de cabeçalho para a solicitação da API
    headers = {
        "Content-type": "application/json",
        "Accept": "application/json",
        "X-VTEX-API-AppKey": "vtexappkey-lojalivreeleve-VNEHUA",
        "X-VTEX-API-AppToken": "LTPWSHLJWLMTZMDBNIMBYXJQOJTCGMWNGGQUNSJYFDEKGUVYLZHKOZJCXCBLIBGUZHNQZOSFLMRQYSMOIPSQXQSPZQXVDISFLRHOVAXXNJFTNSWSBYQPEWOAWQNPYABV"
    }
    
    linhas_processadas = 0
    
    # Iterar sobre todas as linhas da planilha
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        sku_id = row[0]
        if sku_id:
            # Construir a URL substituindo o marcador de posição {skuId}
            api_url = f"https://lojalivreeleve.vtexcommercestable.com.br/api/logistics/pvt/inventory/skus/{sku_id}"

            # Fazer a solicitação para a API
            response = requests.get(api_url, headers=headers)
            if response.status_code == 200:
                total_quantity = response.json()["balance"][0]["totalQuantity"]
                
                # Atualizar o valor na segunda coluna (coluna "IDSKU")
                ws.cell(row=linhas_processadas + 2, column=2, value=total_quantity)
                
                linhas_processadas += 1

                # Salvar as alterações parciais no arquivo Excel
                wb.save("dados_atualizados.xlsx")

                # Aguardar 10 segundos antes da próxima requisição
                time.sleep(5)

    # Salvar as alterações finais no arquivo Excel
    wb.save("dados.xlsx")

if __name__ == "__main__":
    update_stock()
